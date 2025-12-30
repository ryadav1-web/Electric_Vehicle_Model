# drive_cycle.py
import os
import openpyxl

MPH_TO_MPS = 0.44704

# Cycle name -> Excel file (kept next to this .py file)
DRIVE_CYCLE_FILES = {
    "ECE_15": "ECE_15_Cycle.xlsx",
    "EUDC":   "EUDC_Cycle.xlsx",
    "HWFET":  "HWFET_Cycle.xlsx",
    "UDDS": "UDDS_Cycle.xlsx",
    "US06": "US06_Cycle.xlsx"
}

def load_drive_cycle(cycle_name_or_path):
    """
    Returns:
        time_list  : [s]
        speed_list : [m/s]  (Excel is mph, converted here)
    """
    # If user passes a cycle name, map to file next to this script
    key = str(cycle_name_or_path).strip().upper()
    if key in DRIVE_CYCLE_FILES:
        file_path = os.path.join(os.path.dirname(__file__), DRIVE_CYCLE_FILES[key])
    else:
        # otherwise treat it as a direct path
        file_path = str(cycle_name_or_path)

    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb[wb.sheetnames[0]]   # first sheet

    time_list = []
    speed_list = []

    for row in ws.iter_rows(values_only=True):
        t = row[0]
        v = row[1]

        # skip headers / blanks
        if isinstance(t, (int, float)) and isinstance(v, (int, float)):
            time_list.append(float(t))
            speed_list.append(float(v) * MPH_TO_MPS)  # mph -> m/s

    if len(time_list) < 2:
        raise ValueError("Excel must contain at least 2 numeric rows: Time(s), Speed(mph).")

    return time_list, speed_list


def get_vref_interpolated(t, time_list, speed_list):
    """
    Linear interpolation: finds speed at time t.
    """
    if t <= time_list[0]:
        return speed_list[0]
    if t >= time_list[-1]:
        return speed_list[-1]

    # find interval (simple linear scan; fine for most cycles)
    i = 0
    while i < len(time_list) - 1 and time_list[i + 1] < t:
        i += 1

    t1, t2 = time_list[i], time_list[i + 1]
    v1, v2 = speed_list[i], speed_list[i + 1]

    alpha = (t - t1) / (t2 - t1)
    return v1 + alpha * (v2 - v1)
